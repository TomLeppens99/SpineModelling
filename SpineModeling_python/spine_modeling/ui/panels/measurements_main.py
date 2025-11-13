"""
Measurements Main Panel for SpineModeling Application.

This module provides the panel for displaying and managing measurement data
in a table/grid format with database integration.

Translated from: SpineModeling_CSharp/SkeletalModeling/UC_measurementsMain.cs
Original class: UC_measurementsMain

Note: This is a streamlined implementation. Database queries and complex
data operations will be refined during integration testing.
"""

import logging
from typing import Optional, List, Dict, Any
from pathlib import Path
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QHeaderView, QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt

# Set up logging
logger = logging.getLogger(__name__)


class MeasurementsMainPanel(QWidget):
    """
    Panel for displaying and managing measurement data.

    This panel shows measurements in a table format with columns for
    ID, name, comment, and user. Supports filtering, sorting, and
    database operations.

    Attributes:
        measurements_2d_panel: Reference to 2D measurements work panel.
        eos: EOS acquisition data.
        app_data: Application-wide data and settings.
        sql_db: Database connection for measurements.
        subject: Current subject/patient.
        eos_image1: First EOS X-ray image.
        eos_image2: Second EOS X-ray image.
        eos_space: 3D space reconstruction.
        measurement: Current measurement being edited.

    Examples:
        >>> panel = MeasurementsMainPanel()
        >>> panel.refresh_measurements()
    """

    def __init__(self, parent: Optional[QWidget] = None):
        """
        Initialize the measurements main panel.

        Args:
            parent: Optional parent widget.
        """
        super().__init__(parent)

        # References to other components
        self.measurements_2d_panel = None
        self.eos = None
        self.app_data = None
        self.sql_db = None
        self.subject = None
        self.eos_image1 = None
        self.eos_image2 = None
        self.eos_space = None
        self.measurement = None

        # Data
        self._measurements_data: List[Dict[str, Any]] = []

        self._setup_ui()

    def _setup_ui(self) -> None:
        """
        Set up the user interface components.

        Creates a table widget for displaying measurements and toolbar
        for data operations.
        """
        # Main layout
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Toolbar
        toolbar_layout = QHBoxLayout()
        layout.addLayout(toolbar_layout)

        lbl_title = QLabel("Measurements")
        lbl_title.setStyleSheet("font-size: 12pt; font-weight: bold;")
        toolbar_layout.addWidget(lbl_title)

        toolbar_layout.addStretch()

        # Refresh button
        btn_refresh = QPushButton("Refresh")
        btn_refresh.clicked.connect(self.refresh_measurements)
        toolbar_layout.addWidget(btn_refresh)

        # Delete button
        btn_delete = QPushButton("Delete Selected")
        btn_delete.clicked.connect(self._on_delete_selected)
        toolbar_layout.addWidget(btn_delete)

        # Export button
        btn_export = QPushButton("Export to Excel")
        btn_export.clicked.connect(self._on_export_to_excel)
        toolbar_layout.addWidget(btn_export)

        # Export TRC button
        btn_export_trc = QPushButton("Export to TRC")
        btn_export_trc.clicked.connect(self._on_export_to_trc)
        toolbar_layout.addWidget(btn_export_trc)

        # Table widget for measurements
        self.table_measurements = QTableWidget()
        self.table_measurements.setColumnCount(4)
        self.table_measurements.setHorizontalHeaderLabels([
            "Measurement ID",
            "Name",
            "Comment",
            "User"
        ])

        # Set column widths
        header = self.table_measurements.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        # Enable selection
        self.table_measurements.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_measurements.setSelectionMode(QTableWidget.ExtendedSelection)

        # Enable sorting
        self.table_measurements.setSortingEnabled(True)

        layout.addWidget(self.table_measurements)

        # Status label
        self.lbl_status = QLabel("No measurements loaded")
        self.lbl_status.setStyleSheet("color: #666;")
        layout.addWidget(self.lbl_status)

    def refresh_measurements(self, user_id: Optional[int] = None) -> None:
        """
        Refresh measurements from database.

        Loads measurements for the current subject and displays them
        in the table. Optionally filters by user ID.

        Args:
            user_id: Optional user ID to filter measurements.
        """
        # Clear current data
        self.table_measurements.setRowCount(0)
        self._measurements_data.clear()

        if self.sql_db is None:
            self.lbl_status.setText("Database not connected")
            # Load sample data for testing
            self._load_sample_data()
            return

        try:
            # Get all measurements if no specific subject
            if self.subject is None:
                logger.info("No subject selected, loading all measurements")
                measurements = self.sql_db.get_all_measurements()
            else:
                # Get subject from database
                subject_code = self.subject.subject_code if hasattr(self.subject, 'subject_code') else "DEFAULT"
                subject_obj = self.sql_db.get_subject_by_code(subject_code)

                if subject_obj is None:
                    logger.warning(f"Subject {subject_code} not found in database")
                    self.lbl_status.setText(f"Subject {subject_code} not found in database")
                    return

                # Get measurements for this subject
                logger.info(f"Loading measurements for subject {subject_code}")
                measurements = self.sql_db.get_measurements_by_subject(subject_obj.subject_id)

            # Convert measurements to dictionary format for the table
            measurements_data = []
            for meas in measurements:
                meas_dict = {
                    "MeasurementID": meas.measurement_id,
                    "MeasurementName": meas.measurement_name,
                    "MeasurementComment": meas.comment or "",
                    "UserName": meas.user or "Unknown",
                    "MeasurementType": meas.measurement_type,
                    "ImageType": meas.image_type,
                    "MeasurementValue": meas.measurement_value,
                    "MeasurementUnit": meas.measurement_unit,
                    "PosX": meas.x_coord,
                    "PosY": meas.y_coord,
                    "PosZ": meas.z_coord,
                    "EllipseCenterX": meas.ellipse_center_x,
                    "EllipseCenterY": meas.ellipse_center_y,
                    "EllipseMajorAxis": meas.ellipse_major_axis,
                    "EllipseMinorAxis": meas.ellipse_minor_axis,
                    "EllipseAngle": meas.ellipse_angle,
                    "MeasurementDate": meas.measurement_date
                }
                measurements_data.append(meas_dict)

            # Populate table
            self._populate_table(measurements_data)

            # Update status
            count = len(measurements_data)
            subject_info = f" for subject {self.subject.subject_code}" if self.subject else ""
            self.lbl_status.setText(f"Loaded {count} measurement(s){subject_info}")
            logger.info(f"Loaded {count} measurements from database")

        except Exception as e:
            logger.error(f"Error loading measurements from database: {e}", exc_info=True)
            QMessageBox.critical(
                self,
                "Database Error",
                f"Failed to load measurements:\n{e}"
            )
            self.lbl_status.setText(f"Error: {e}")

    def _load_sample_data(self) -> None:
        """
        Load sample measurement data for testing.

        Creates dummy measurement records to demonstrate the table.
        """
        sample_data = [
            {
                "MeasurementID": 1,
                "MeasurementName": "L1 Superior Endplate",
                "MeasurementComment": "Lumbar vertebra L1",
                "UserName": "Dr. Smith"
            },
            {
                "MeasurementID": 2,
                "MeasurementName": "L2 Superior Endplate",
                "MeasurementComment": "Lumbar vertebra L2",
                "UserName": "Dr. Smith"
            },
            {
                "MeasurementID": 3,
                "MeasurementName": "L3 Superior Endplate",
                "MeasurementComment": "Lumbar vertebra L3",
                "UserName": "Dr. Jones"
            },
            {
                "MeasurementID": 4,
                "MeasurementName": "Sacrum Marker",
                "MeasurementComment": "S1 landmark",
                "UserName": "Dr. Smith"
            },
        ]

        self._populate_table(sample_data)
        self.lbl_status.setText(f"Loaded {len(sample_data)} measurements (sample data)")

    def _populate_table(self, data: List[Dict[str, Any]]) -> None:
        """
        Populate the table with measurement data.

        Args:
            data: List of measurement dictionaries with keys:
                  MeasurementID, MeasurementName, MeasurementComment, UserName
        """
        self._measurements_data = data
        self.table_measurements.setRowCount(len(data))

        for row_idx, measurement in enumerate(data):
            # Measurement ID
            item_id = QTableWidgetItem(str(measurement.get("MeasurementID", "")))
            item_id.setFlags(item_id.flags() & ~Qt.ItemIsEditable)  # Read-only
            self.table_measurements.setItem(row_idx, 0, item_id)

            # Name
            item_name = QTableWidgetItem(measurement.get("MeasurementName", ""))
            self.table_measurements.setItem(row_idx, 1, item_name)

            # Comment
            item_comment = QTableWidgetItem(measurement.get("MeasurementComment", ""))
            self.table_measurements.setItem(row_idx, 2, item_comment)

            # User
            item_user = QTableWidgetItem(measurement.get("UserName", ""))
            item_user.setFlags(item_user.flags() & ~Qt.ItemIsEditable)  # Read-only
            self.table_measurements.setItem(row_idx, 3, item_user)

    def _on_delete_selected(self) -> None:
        """
        Handle Delete Selected button click.

        Deletes the selected measurements from the database after
        confirmation.
        """
        selected_rows = self.table_measurements.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.information(
                self,
                "No Selection",
                "Please select one or more measurements to delete."
            )
            return

        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Deletion",
            f"Are you sure you want to delete {len(selected_rows)} measurement(s)?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.sql_db is None:
                QMessageBox.warning(
                    self,
                    "Database Error",
                    "Database not connected. Cannot delete measurements."
                )
                return

            try:
                # Get measurement IDs to delete
                measurement_ids = []
                for row_index in selected_rows:
                    row = row_index.row()
                    item = self.table_measurements.item(row, 0)  # Column 0 is ID
                    if item:
                        try:
                            meas_id = int(item.text())
                            measurement_ids.append(meas_id)
                        except ValueError:
                            pass

                # Delete from database
                deleted_count = 0
                for meas_id in measurement_ids:
                    if self.sql_db.delete_measurement(meas_id):
                        deleted_count += 1
                        logger.info(f"Deleted measurement ID {meas_id}")
                    else:
                        logger.warning(f"Failed to delete measurement ID {meas_id}")

                # Refresh the table
                self.refresh_measurements()

                # Show success message
                self.lbl_status.setText(f"Deleted {deleted_count} measurement(s)")
                if deleted_count > 0:
                    QMessageBox.information(
                        self,
                        "Success",
                        f"Successfully deleted {deleted_count} measurement(s)."
                    )

            except Exception as e:
                logger.error(f"Error deleting measurements: {e}", exc_info=True)
                QMessageBox.critical(
                    self,
                    "Database Error",
                    f"Failed to delete measurements:\n{e}"
                )
                self.lbl_status.setText(f"Error deleting measurements")

    def _on_export_to_excel(self) -> None:
        """
        Handle Export to Excel button click.

        Exports the current measurements to an Excel file using openpyxl.
        Based on C# ExportToExcel() method in UC_measurementsMain.cs
        """
        if self.table_measurements.rowCount() == 0:
            QMessageBox.information(
                self,
                "No Data",
                "No measurements to export."
            )
            return

        # Show file save dialog
        default_filename = f"measurements_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Measurements to Excel",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return  # User cancelled

        try:
            self.export_to_excel(file_path)
            QMessageBox.information(
                self,
                "Export Successful",
                f"Measurements exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export measurements:\n{e}"
            )

    def export_to_excel(self, file_path: str) -> None:
        """
        Export measurements table to Excel file.

        Uses openpyxl to create a formatted Excel workbook with measurement data.
        Exports all columns: MeasurementID, Name, Comment, User, and additional
        3D coordinates if available.

        Args:
            file_path: Output file path for the Excel file.

        Raises:
            ImportError: If openpyxl is not installed.
            Exception: If export fails.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Measurements Export"

        # Add header row with formatting
        headers = ["Measurement ID", "Name", "Comment", "User"]

        # Add 3D coordinate headers if we have that data
        if self._measurements_data and any(
            'PosX' in m or 'PosY' in m or 'PosZ' in m
            for m in self._measurements_data
        ):
            headers.extend(["Pos X (m)", "Pos Y (m)", "Pos Z (m)"])

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_idx in range(self.table_measurements.rowCount()):
            # Get data from table
            meas_id = self.table_measurements.item(row_idx, 0)
            name = self.table_measurements.item(row_idx, 1)
            comment = self.table_measurements.item(row_idx, 2)
            user = self.table_measurements.item(row_idx, 3)

            # Write to Excel (row_idx + 2 because Excel is 1-indexed and row 1 is headers)
            excel_row = row_idx + 2
            ws.cell(row=excel_row, column=1, value=meas_id.text() if meas_id else "")
            ws.cell(row=excel_row, column=2, value=name.text() if name else "")
            ws.cell(row=excel_row, column=3, value=comment.text() if comment else "")
            ws.cell(row=excel_row, column=4, value=user.text() if user else "")

            # Add 3D coordinates if available
            if row_idx < len(self._measurements_data):
                measurement = self._measurements_data[row_idx]
                if 'PosX' in measurement:
                    ws.cell(row=excel_row, column=5, value=measurement.get('PosX', ''))
                if 'PosY' in measurement:
                    ws.cell(row=excel_row, column=6, value=measurement.get('PosY', ''))
                if 'PosZ' in measurement:
                    ws.cell(row=excel_row, column=7, value=measurement.get('PosZ', ''))

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(file_path)
        self.lbl_status.setText(f"Exported {self.table_measurements.rowCount()} measurements to Excel")

    def _on_export_to_trc(self) -> None:
        """
        Handle Export to TRC button click.

        Exports selected measurements to TRC (Track Row Column) format,
        which is a motion capture marker trajectory file format used by OpenSim.
        """
        selected_rows = self.table_measurements.selectionModel().selectedRows()

        if not selected_rows:
            QMessageBox.information(
                self,
                "No Selection",
                "Please select one or more measurements to export to TRC format."
            )
            return

        # Show file save dialog
        default_filename = f"marker_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.trc"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Markers to TRC",
            default_filename,
            "TRC Files (*.trc);;All Files (*)"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Get selected measurements
            selected_measurements = []
            for row_index in selected_rows:
                row = row_index.row()
                if row < len(self._measurements_data):
                    selected_measurements.append(self._measurements_data[row])

            self.export_to_trc(file_path, selected_measurements)
            QMessageBox.information(
                self,
                "Export Successful",
                f"Exported {len(selected_measurements)} markers to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export to TRC:\n{e}"
            )

    def export_to_trc(self, file_path: str, measurements: List[Dict[str, Any]]) -> None:
        """
        Export measurements to TRC (Track Row Column) format.

        TRC is a tab-delimited text file format used by OpenSim for marker trajectories.
        Format based on C# PrintMarkersToTrc() method in UC_measurementsMain.cs

        Args:
            file_path: Output file path for the TRC file.
            measurements: List of measurement dictionaries with PosX, PosY, PosZ coordinates.

        Raises:
            Exception: If export fails.
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for TRC export. "
                "Install it with: pip install openpyxl"
            )

        # Create workbook (we'll export to Excel first, then convert to tab-delimited)
        wb = Workbook()
        ws = wb.active

        # TRC file header - Row 1
        ws.cell(1, 1, "PathFileType")
        ws.cell(1, 2, 4)  # Version
        ws.cell(1, 3, "(X/Y/Z)")
        ws.cell(1, 4, Path(file_path).name)

        # TRC file metadata - Row 2
        ws.cell(2, 1, "DataRate")
        ws.cell(2, 2, "CameraRate")
        ws.cell(2, 3, "NumFrames")
        ws.cell(2, 4, "NumMarkers")
        ws.cell(2, 5, "Units")
        ws.cell(2, 6, "OrigDataRate")
        ws.cell(2, 7, "OrigDataStartFrame")
        ws.cell(2, 8, "OrigNumFrames")

        # TRC metadata values - Row 3
        num_markers = len(measurements)
        num_frames = 60  # Static pose, but TRC requires frame data
        camera_rate = 60

        ws.cell(3, 1, camera_rate)
        ws.cell(3, 2, camera_rate)
        ws.cell(3, 3, num_frames)
        ws.cell(3, 4, num_markers)
        ws.cell(3, 5, "mm")  # Units in millimeters
        ws.cell(3, 6, camera_rate)
        ws.cell(3, 7, 1)  # Start frame
        ws.cell(3, 8, num_frames)

        # Column headers - Row 4
        ws.cell(4, 1, "Frame#")
        ws.cell(4, 2, "Time")

        # Marker names - Row 4
        for i, measurement in enumerate(measurements):
            marker_name = measurement.get("MeasurementName", f"Marker{i+1}")
            ws.cell(4, 3 * (i + 1), marker_name)

        # X/Y/Z labels - Row 5
        ws.cell(5, 1, "")
        ws.cell(5, 2, "")
        for i in range(num_markers):
            ws.cell(5, 3 * (i + 1), f"X{i+1}")
            ws.cell(5, (3 * (i + 1)) + 1, f"Y{i+1}")
            ws.cell(5, (3 * (i + 1)) + 2, f"Z{i+1}")

        # Frame data - Rows 6 onwards
        # For static markers, repeat the same position for all frames
        for frame in range(1, num_frames + 1):
            row_num = 6 + frame - 1

            # Frame number
            ws.cell(row_num, 1, frame)

            # Time (in seconds)
            time = round((1 / camera_rate) * (frame - 1), 3)
            ws.cell(row_num, 2, time)

            # Marker coordinates (convert from meters to millimeters)
            for i, measurement in enumerate(measurements):
                pos_x = measurement.get("PosX", 0.0)
                pos_y = measurement.get("PosY", 0.0)
                pos_z = measurement.get("PosZ", 0.0)

                # Convert meters to millimeters and format with decimal point
                ws.cell(row_num, 3 * (i + 1), pos_x * 1000)
                ws.cell(row_num, (3 * (i + 1)) + 1, pos_y * 1000)
                ws.cell(row_num, (3 * (i + 1)) + 2, pos_z * 1000)

        # Save as Excel first
        temp_excel = file_path + ".temp.xlsx"
        wb.save(temp_excel)

        # Convert to tab-delimited text (TRC format)
        wb_read = Workbook()
        from openpyxl import load_workbook
        wb_read = load_workbook(temp_excel)
        ws_read = wb_read.active

        with open(file_path, 'w', encoding='utf-8') as f:
            for row in ws_read.iter_rows():
                values = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        values.append('')
                    else:
                        # Format floats with decimal point (not comma)
                        if isinstance(value, float):
                            values.append(str(value).replace(',', '.'))
                        else:
                            values.append(str(value))
                # Join with tabs
                f.write('\t'.join(values) + '\n')

        # Clean up temp file
        import os
        os.remove(temp_excel)

        self.lbl_status.setText(f"Exported {num_markers} markers to TRC format")

    def get_selected_measurement_ids(self) -> List[int]:
        """
        Get the IDs of selected measurements.

        Returns:
            List of measurement IDs for selected rows.
        """
        selected_rows = self.table_measurements.selectionModel().selectedRows()
        ids = []

        for row in selected_rows:
            item = self.table_measurements.item(row.row(), 0)  # Column 0 is ID
            if item:
                try:
                    ids.append(int(item.text()))
                except ValueError:
                    pass

        return ids
